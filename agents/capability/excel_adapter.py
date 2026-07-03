import openpyxl
from orchestrator.schemas import CapabilityCheckInput, CapabilityCheckResult, CapabilityType

class ExcelAdapter:
    """
    Phase 15: Validates Excel workbooks (Read-Only).
    """
    capability_type: CapabilityType = CapabilityType.EXCEL

    def run(self, payload: CapabilityCheckInput) -> CapabilityCheckResult:
        params = payload.params
        expected = payload.expected or {}
        file_path = params.get("file_path")
        sheet_name = params.get("sheet_name")
        
        if not file_path:
            return self._fail("Missing 'file_path'")
            
        wb = None
        try:
            # Debug Fix: Removed read_only=True to allow ws["A1"] string coordinate access
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return CapabilityCheckResult(
                        capability=self.capability_type, passed=False, confidence=1.0,
                        evidence={"error": f"Sheet '{sheet_name}' not found", "available_sheets": wb.sheetnames}, escalate=False
                    )
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title
                
            evidence = {"sheet_name": sheet_name, "sheet_exists": True}
            passed = True
            
            # Handle case where max_row might be None for completely empty sheets
            row_count = ws.max_row or 0
            evidence["row_count"] = row_count
            if expected.get("row_count") is not None and row_count != expected["row_count"]:
                passed = False
                evidence["row_count_mismatch"] = True
                
            if expected.get("cell_values"):
                cell_mismatches = {}
                for coord, expected_val in expected["cell_values"].items():
                    actual_val = ws[coord].value
                    if actual_val != expected_val:
                        passed = False
                        cell_mismatches[coord] = {"expected": expected_val, "actual": actual_val}
                if cell_mismatches:
                    evidence["cell_mismatches"] = cell_mismatches
                    
            return CapabilityCheckResult(
                capability=self.capability_type, passed=passed, 
                confidence=1.0 if passed else 0.5, evidence=evidence, escalate=False
            )
        except Exception as e:
            return self._fail(f"Excel processing error: {str(e)}")
        finally:
            if wb: wb.close()

    def _fail(self, msg):
        return CapabilityCheckResult(
            capability=self.capability_type, passed=False, confidence=1.0,
            evidence={"error": msg}, escalate=False
        )