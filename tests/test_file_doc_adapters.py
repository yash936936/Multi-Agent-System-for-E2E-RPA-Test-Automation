from unittest.mock import patch, MagicMock
from orchestrator.schemas import CapabilityCheckInput, CapabilityType
from agents.capability.file_adapter import FileAdapter
from agents.capability.excel_adapter import ExcelAdapter
from agents.capability.pdf_adapter import PdfAdapter

# --- File Adapter Tests ---
def test_file_adapter_local_stat_success(tmp_path):
    test_file = tmp_path / "report.txt"
    test_file.write_text("A" * 100) # 100 bytes
    
    adapter = FileAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.FILE_SYSTEM,
        target=str(test_file),
        params={"action": "local_stat", "path": str(test_file)},
        expected={"exists": True, "min_size_bytes": 50}
    )
    result = adapter.run(payload)
    assert result.passed is True
    assert result.evidence["size_bytes"] == 100

def test_file_adapter_local_hash_success(tmp_path):
    test_file = tmp_path / "data.bin"
    test_file.write_bytes(b"test_data")
    expected_hash = "e7d87b738825c33824cf3fd32b7314161fc8c425129163ff5e7260fc7288da36"  # sha256 of test_data
    
    adapter = FileAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.FILE_SYSTEM, target="",
        params={"action": "local_hash", "path": str(test_file), "hash_algorithm": "sha256"},
        expected={"hash": expected_hash}
    )
    result = adapter.run(payload)
    assert result.passed is True

# --- Excel Adapter Tests ---
def test_excel_adapter_cell_values(tmp_path):
    import openpyxl
    test_file = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["B2"] = 42
    wb.save(test_file)
    
    adapter = ExcelAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.EXCEL, target=str(test_file),
        params={"file_path": str(test_file)},
        expected={"cell_values": {"A1": "Header", "B2": 42}}
    )
    result = adapter.run(payload)
    assert result.passed is True

def test_excel_adapter_sheet_not_found(tmp_path):
    import openpyxl
    test_file = tmp_path / "test2.xlsx"
    openpyxl.Workbook().save(test_file)
    
    adapter = ExcelAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.EXCEL, target=str(test_file),
        params={"file_path": str(test_file), "sheet_name": "MissingSheet"},
        expected={}
    )
    result = adapter.run(payload)
    assert result.passed is False
    assert "not found" in result.evidence["error"]

# --- PDF Adapter Tests ---
def test_pdf_adapter_text_contains(tmp_path):
    # Mocking pypdf to avoid needing a real PDF generator in tests
    adapter = PdfAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.PDF_OCR, target="dummy.pdf",
        params={"file_path": "dummy.pdf"},
        expected={"page_count": 2, "text_contains": ["AURA", "QA"]}
    )
    
    with patch("agents.capability.pdf_adapter.PdfReader") as mock_reader:
        mock_instance = MagicMock()
        mock_reader.return_value = mock_instance
        mock_instance.pages = [MagicMock(), MagicMock()]
        mock_instance.pages[0].extract_text.return_value = "Welcome to AURA"
        mock_instance.pages[1].extract_text.return_value = "This is a QA test"
        mock_instance.is_encrypted = False
        mock_instance.metadata = {}
        
        result = adapter.run(payload)
        assert result.passed is True
        assert result.evidence["page_count"] == 2