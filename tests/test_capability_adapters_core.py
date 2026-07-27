"""Merged test file: test_capability_adapters_core.py
Consolidated from: test_db_adapter.py, test_db_seed_adapter.py, test_core_adapters.py, test_file_doc_adapters.py
All original test functions preserved 1:1. Colliding fixture/helper names
renamed with a source-file suffix to avoid silent shadowing across sections.
"""
from __future__ import annotations
from agents.capability.db_adapter import DbAdapter
from orchestrator.schemas import CapabilityCheckInput, CapabilityType
import sqlite3
from unittest.mock import patch
import pytest
import sqlalchemy
from agents.capability.db_seed_adapter import DbSeedAdapter
from config.settings import settings
from unittest.mock import patch, MagicMock
from agents.capability.api_adapter import ApiAdapter
from agents.capability.email_adapter import EmailAdapter
from agents.capability.file_adapter import FileAdapter
from agents.capability.excel_adapter import ExcelAdapter
from agents.capability.pdf_adapter import PdfAdapter


# ============================================================================
# ---- from test_db_adapter.py ----
# ============================================================================
def _run(query: str, connection_string: str = "sqlite:///:memory:", expected: dict | None = None):
    adapter = DbAdapter()
    return adapter.run(
        CapabilityCheckInput(
            capability=CapabilityType.DATABASE,
            target="",
            params={"connection_string": connection_string, "query": query},
            expected=expected or {},
        )
    )


def test_select_query_still_works():
    result = _run("SELECT 1 as one", expected={"row_count": 1})
    assert result.passed is True
    assert result.evidence["row_count"] == 1


def test_with_cte_query_still_works():
    result = _run("WITH x AS (SELECT 1 as v) SELECT * FROM x")
    assert result.passed is True


def test_drop_table_is_refused():
    result = _run("DROP TABLE users")
    assert result.passed is False
    assert "read-only" in result.evidence["error"].lower()


def test_delete_is_refused():
    result = _run("DELETE FROM users WHERE 1=1")
    assert result.passed is False


def test_update_is_refused():
    result = _run("UPDATE users SET role='admin' WHERE username='attacker'")
    assert result.passed is False


def test_insert_is_refused():
    result = _run("INSERT INTO users (username, role) VALUES ('x', 'admin')")
    assert result.passed is False


def test_stacked_statement_smuggling_is_rejected_by_the_driver():
    # A SELECT prefix alone wouldn't stop a stacked "SELECT 1; DROP
    # TABLE ..." -- but SQLAlchemy's default execute() doesn't support
    # multiple statements per call at all, so this fails safely with a
    # driver-level error before the DROP ever runs. Documented here so
    # this protection isn't accidentally assumed to come from the
    # allowlist above if the driver/execution style ever changes.
    result = _run("SELECT 1; DROP TABLE users;--")
    assert result.passed is False


def test_missing_query_still_fails_clearly():
    result = _run("")
    assert result.passed is False


def test_mutating_function_inside_select_is_refused():
    # 2026-07-13 (decisions.md D-017 / roadmap issue 1.7): the prefix
    # allowlist alone lets a syntactically-SELECT query through even if it
    # calls a mutating/exfiltration function. These should now be refused
    # by the secondary denylist check.
    dangerous_queries = [
        "SELECT setval('users_id_seq', 1)",
        "SELECT pg_terminate_backend(123)",
        "SELECT lo_export(16420, '/tmp/x')",
        "SELECT LOAD_FILE('/etc/passwd')",
        "SELECT * FROM t INTO OUTFILE '/tmp/x.csv'",
        "SELECT dblink_exec('conn', 'DROP TABLE users')",
        "SELECT * FROM OPENROWSET('SQLNCLI', 'evil')",
    ]
    for q in dangerous_queries:
        result = _run(q)
        assert result.passed is False, f"expected refusal for: {q}"
        assert "error" in result.evidence


def test_legitimate_select_queries_are_not_false_flagged():
    # Guard against the new denylist being over-broad and breaking real
    # read-only assertions that happen to mention similar-looking words
    # (e.g. a column literally named "execution_id").
    safe_queries = [
        "SELECT 1 as execution_id, 'ok' as status",
        "SELECT name FROM sqlite_master WHERE type = 'table'",
        "WITH recent AS (SELECT 1 as v) SELECT * FROM recent",
    ]
    for q in safe_queries:
        result = _run(q)
        assert result.passed is True, f"expected pass for safe query: {q} -- got {result.evidence}"


def test_query_error_healing_hints_include_exception_text():
    # 2026-07-13 (decisions.md D-017 / roadmap issue 1.6): healing_hints
    # used to omit the "exception" key entirely, so
    # cross_modal_diagnoser.py's column-drift regex always matched an
    # empty string and could never actually detect a "column does not
    # exist" error. This asserts the real driver error text now reaches
    # healing_hints, and that CrossModalDiagnoser's regex can actually
    # match it end-to-end.
    result = _run("SELECT nonexistent_column_xyz FROM sqlite_master")
    assert result.passed is False
    assert "healing_hints" in result.evidence
    hints = result.evidence["healing_hints"]
    assert "exception" in hints
    assert hints["exception"] == result.evidence["exception"]

    from agents.planner.cross_modal_diagnoser import CrossModalDiagnoser
    from orchestrator.schemas import TestStep, CapabilityType as CT

    diagnoser = CrossModalDiagnoser()
    step = TestStep(
        step_id=1, action="capability_check", capability_type=CT.DATABASE,
        params={}, expected={},
    )
    # sqlite's real error text doesn't match the Postgres-flavored
    # "column X does not exist" pattern this heuristic targets, so this
    # confirms the data now *reaches* the regex (no crash, real text
    # passed through) even though sqlite's message won't match it --
    # a Postgres-backed run would. The important fix is that hints.get(
    # "exception") is no longer silently empty.
    result_step = diagnoser.diagnose(step, result)
    assert result_step is None  # no safe rename target either way; escalate

# ============================================================================
# ---- from test_db_seed_adapter.py ----
# ============================================================================
@pytest.fixture
def seeding_enabled():
    original = settings.allow_db_seeding
    settings.allow_db_seeding = True
    yield
    settings.allow_db_seeding = original


@pytest.fixture
def sqlite_db(tmp_path):
    db_path = tmp_path / "seed_test.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("CREATE TABLE users (id INTEGER PRIMARY KEY, username TEXT, role TEXT)")
    conn.commit()
    conn.close()
    return f"sqlite:///{db_path}"


def _run__db_seed_adapter(connection_string, table=None, values=None, rows=None, capability_params_extra=None):
    adapter = DbSeedAdapter()
    params = {"connection_string": connection_string}
    if table is not None:
        params["table"] = table
    if values is not None:
        params["values"] = values
    if rows is not None:
        params["rows"] = rows
    if capability_params_extra:
        params.update(capability_params_extra)
    return adapter.run(
        CapabilityCheckInput(
            capability=CapabilityType.DB_SEED,
            target="",
            params=params,
            expected={},
        )
    )


def _select_all(connection_string, table):
    engine = sqlalchemy.create_engine(connection_string)
    with engine.connect() as conn:
        result = conn.execute(sqlalchemy.text(f"SELECT * FROM {table}"))
        return [dict(zip(result.keys(), row)) for row in result.fetchall()]


# --- Gate ---

def test_disabled_by_default_refuses_even_with_valid_params(sqlite_db):
    assert settings.allow_db_seeding is False
    result = _run__db_seed_adapter(sqlite_db, table="users", values={"id": 1, "username": "alice", "role": "admin"})
    assert result.passed is False
    assert "allow_db_seeding" in result.evidence["error"]
    # Confirm nothing was actually written.
    assert _select_all(sqlite_db, "users") == []


def test_enabled_allows_seeding(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="users", values={"id": 1, "username": "alice", "role": "admin"})
    assert result.passed is True
    rows = _select_all(sqlite_db, "users")
    assert rows == [{"id": 1, "username": "alice", "role": "admin"}]


# --- Single row / batch rows ---

def test_single_row_via_values(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="users", values={"id": 1, "username": "bob", "role": "user"})
    assert result.passed is True
    assert result.evidence["row_count"] == 1


def test_batch_rows(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(
        sqlite_db,
        table="users",
        rows=[
            {"id": 1, "username": "alice", "role": "admin"},
            {"id": 2, "username": "bob", "role": "user"},
        ],
    )
    assert result.passed is True
    assert result.evidence["row_count"] == 2
    rows = _select_all(sqlite_db, "users")
    assert len(rows) == 2


def test_mismatched_row_shapes_rejected(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(
        sqlite_db,
        table="users",
        rows=[
            {"id": 1, "username": "alice", "role": "admin"},
            {"id": 2, "username": "bob"},  # missing 'role'
        ],
    )
    assert result.passed is False
    assert "same columns" in result.evidence["error"]
    assert _select_all(sqlite_db, "users") == []


def test_empty_rows_list_rejected(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="users", rows=[])
    assert result.passed is False
    assert "empty" in result.evidence["error"].lower()


def test_missing_values_and_rows_rejected(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="users")
    assert result.passed is False
    assert "values" in result.evidence["error"] or "rows" in result.evidence["error"]


# --- Identifier validation (table/column names are interpolated, not bound) ---

def test_invalid_table_name_rejected(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="users; DROP TABLE users; --", values={"id": 1})
    assert result.passed is False
    assert "table name" in result.evidence["error"].lower()


def test_invalid_column_name_rejected(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="users", values={"id; DROP TABLE users; --": 1})
    assert result.passed is False
    assert "column name" in result.evidence["error"].lower()


def test_valid_snake_case_identifiers_pass(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="users", values={"id": 3, "username": "carol", "role": "admin"})
    assert result.passed is True


# --- Only INSERT is possible: there's no query-string param to abuse ---

def test_no_query_param_accepted_at_all(seeding_enabled, sqlite_db):
    """Even if a caller tries to sneak a 'query' param in (as if this were
    db_adapter.py), it's simply ignored -- there is no code path in this
    adapter that reads or executes arbitrary SQL text."""
    result = _run__db_seed_adapter(
        sqlite_db,
        table="users",
        values={"id": 1, "username": "x", "role": "y"},
        capability_params_extra={"query": "DROP TABLE users"},
    )
    assert result.passed is True
    # Table still exists and has exactly the one seeded row -- the
    # smuggled 'query' param had no effect.
    assert len(_select_all(sqlite_db, "users")) == 1


# --- Audit logging ---

def test_seed_operation_is_audited_with_exact_rows(seeding_enabled, sqlite_db):
    with patch("agents.capability.db_seed_adapter.audit_logger") as mock_audit:
        result = _run__db_seed_adapter(
            sqlite_db,
            table="users",
            rows=[{"id": 1, "username": "alice", "role": "admin"}],
        )
    assert result.passed is True
    mock_audit.log.assert_called_once()
    _, kwargs = mock_audit.log.call_args
    assert kwargs["action"] == "DB_SEED"
    assert kwargs["resource"] == "users"
    assert kwargs["details"]["rows"] == [{"id": 1, "username": "alice", "role": "admin"}]
    assert kwargs["details"]["row_count"] == 1


def test_failed_seed_is_not_audited(seeding_enabled, sqlite_db):
    with patch("agents.capability.db_seed_adapter.audit_logger") as mock_audit:
        result = _run__db_seed_adapter(sqlite_db, table="users", rows=[])
    assert result.passed is False
    mock_audit.log.assert_not_called()


# --- DB-level failure surfaces cleanly (e.g. table doesn't exist) ---

def test_nonexistent_table_fails_cleanly(seeding_enabled, sqlite_db):
    result = _run__db_seed_adapter(sqlite_db, table="does_not_exist", values={"id": 1})
    assert result.passed is False
    assert result.escalate is True
    assert "exception" in result.evidence

# ============================================================================
# ---- from test_core_adapters.py ----
# ============================================================================
@pytest.fixture
def api_payload():
    return CapabilityCheckInput(
        capability=CapabilityType.API,
        target="https://api.test.com/1",
        params={"method": "GET", "url": "https://api.test.com/1"},
        expected={"status": 200, "json": {"id": 1}}
    )

@pytest.fixture
def db_payload():
    return CapabilityCheckInput(
        capability=CapabilityType.DATABASE,
        target="sqlite:///:memory:",
        params={"connection_string": "sqlite:///:memory:", "query": "SELECT 1 as id, 'test' as name"},
        expected={"row_count": 1, "values": {"id": 1, "name": "test"}}
    )

@pytest.fixture
def email_payload():
    return CapabilityCheckInput(
        capability=CapabilityType.EMAIL,
        target="imap.test.com",
        params={"action": "poll", "imap_server": "imap.test.com", "username": "u", "password": "p"},
        expected={"subject": "Test"}
    )

# --- API Adapter Tests ---
def test_api_adapter_success(api_payload):
    adapter = ApiAdapter()
    with patch("httpx.Client") as mock_client:
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.elapsed.total_seconds.return_value = 0.1
        mock_response.json.return_value = {"id": 1, "name": "Test"}
        mock_client.return_value.__enter__.return_value.request.return_value = mock_response
        
        result = adapter.run(api_payload)
        assert result.passed is True
        assert result.evidence["status_code"] == 200

def test_api_adapter_missing_url():
    adapter = ApiAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.API, target="", params={}, expected={}
    )
    result = adapter.run(payload)
    assert result.passed is False
    assert "Missing 'url'" in result.evidence.get("error", "")

# --- DB Adapter Tests ---
def test_db_adapter_success(db_payload):
    adapter = DbAdapter()
    result = adapter.run(db_payload)
    assert result.passed is True
    assert result.evidence["row_count"] == 1

def test_db_adapter_failure_values():
    adapter = DbAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.DATABASE,
        target="sqlite:///:memory:",
        params={"connection_string": "sqlite:///:memory:", "query": "SELECT 1 as id, 'wrong' as name"},
        expected={"values": {"name": "test"}}
    )
    result = adapter.run(payload)
    assert result.passed is False
    assert "value_mismatch_name" in result.evidence

# --- Email Adapter Tests ---
def test_email_adapter_poll_success(email_payload):
    adapter = EmailAdapter()
    with patch("imaplib.IMAP4_SSL") as mock_imap:
        mock_mail = MagicMock()
        mock_imap.return_value = mock_mail
        mock_mail.login.return_value = ("OK", [b""])
        mock_mail.select.return_value = ("OK", [b"1"])
        mock_mail.search.return_value = ("OK", [b"1"])
        
        raw_email = b"From: sender@test.com\r\nSubject: Test Email\r\n\r\nBody text"
        mock_mail.fetch.return_value = ("OK", [(b"1", raw_email)])
        
        result = adapter.run(email_payload)
        assert result.passed is True

def test_email_adapter_send_success():
    adapter = EmailAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.EMAIL,
        target="smtp.test.com",
        params={"action": "send", "smtp_server": "smtp.test.com", "username": "u", "password": "p", "to": "r@test.com"},
        expected={}
    )
    with patch("smtplib.SMTP") as mock_smtp:
        mock_server = MagicMock()
        mock_smtp.return_value.__enter__.return_value = mock_server
        result = adapter.run(payload)
        assert result.passed is True

# ============================================================================
# ---- from test_file_doc_adapters.py ----
# ============================================================================
# --- File Adapter Tests ---
def test_file_adapter_local_stat_success(tmp_path):
    test_file = tmp_path / "report.txt"
    test_file.write_text("A" * 100) # 100 bytes
    
    adapter = FileAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.FILE_SYSTEM,
        target=str(test_file),
        params={"action": "local_stat", "path": str(test_file)},
        expected={"exists": True, "min_size_bytes": 50}
    )
    result = adapter.run(payload)
    assert result.passed is True
    assert result.evidence["size_bytes"] == 100

def test_file_adapter_local_hash_success(tmp_path):
    test_file = tmp_path / "data.bin"
    test_file.write_bytes(b"test_data")
    expected_hash = "e7d87b738825c33824cf3fd32b7314161fc8c425129163ff5e7260fc7288da36"  # sha256 of test_data
    
    adapter = FileAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.FILE_SYSTEM, target="",
        params={"action": "local_hash", "path": str(test_file), "hash_algorithm": "sha256"},
        expected={"hash": expected_hash}
    )
    result = adapter.run(payload)
    assert result.passed is True

# --- Excel Adapter Tests ---
def test_excel_adapter_cell_values(tmp_path):
    import openpyxl
    test_file = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["B2"] = 42
    wb.save(test_file)
    
    adapter = ExcelAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.EXCEL, target=str(test_file),
        params={"file_path": str(test_file)},
        expected={"cell_values": {"A1": "Header", "B2": 42}}
    )
    result = adapter.run(payload)
    assert result.passed is True

def test_excel_adapter_sheet_not_found(tmp_path):
    import openpyxl
    test_file = tmp_path / "test2.xlsx"
    openpyxl.Workbook().save(test_file)
    
    adapter = ExcelAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.EXCEL, target=str(test_file),
        params={"file_path": str(test_file), "sheet_name": "MissingSheet"},
        expected={}
    )
    result = adapter.run(payload)
    assert result.passed is False
    assert "not found" in result.evidence["error"]

# --- PDF Adapter Tests ---
def test_pdf_adapter_text_contains(tmp_path):
    # Mocking pypdf to avoid needing a real PDF generator in tests
    adapter = PdfAdapter()
    payload = CapabilityCheckInput(
        capability=CapabilityType.PDF_OCR, target="dummy.pdf",
        params={"file_path": "dummy.pdf"},
        expected={"page_count": 2, "text_contains": ["AURA", "QA"]}
    )
    
    with patch("agents.capability.pdf_adapter.PdfReader") as mock_reader:
        mock_instance = MagicMock()
        mock_reader.return_value = mock_instance
        mock_instance.pages = [MagicMock(), MagicMock()]
        mock_instance.pages[0].extract_text.return_value = "Welcome to AURA"
        mock_instance.pages[1].extract_text.return_value = "This is a QA test"
        mock_instance.is_encrypted = False
        mock_instance.metadata = {}
        
        result = adapter.run(payload)
        assert result.passed is True
        assert result.evidence["page_count"] == 2
